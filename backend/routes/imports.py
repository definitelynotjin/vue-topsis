from flask import Blueprint, request, jsonify
from db import db
from models import Alternative, Criteria, Score
import openpyxl
from io import BytesIO

import_bp = Blueprint("import", __name__)

@import_bp.route("/import/<int:project_id>", methods=["POST"])
def import_alternatives(project_id):
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "Empty filename"}), 400

    # Baca Excel
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # Ambil header dari baris pertama
    headers = [cell.value for cell in ws[1]]

    # Cek kolom wajib
    required_cols = ["Nama", "ID"]
    for col in required_cols:
        if col not in headers:
            return jsonify({"error": f"Missing column: {col}"}), 400

    # Ambil kriteria untuk project
    criteria_list = Criteria.query.filter_by(project_id=project_id).all()
    criteria_map = {c.name: c for c in criteria_list}

    # Cek semua kriteria ada di header Excel
    for cname in criteria_map.keys():
        if cname not in headers:
            return jsonify({"error": f"Criteria '{cname}' not found int File"}), 400

    # Index kolom
    col_index = {name: headers.index(name) for name in headers}

    imported = []

    # Iterasi data mulai baris kedua
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_alt = row[col_index["Nama"]]
        id_alt = row[col_index["ID"]]

        # Buat Alternative
        alt = Alternative(name=name_alt, id_alt=id_alt, project_id=project_id)
        db.session.add(alt)
        db.session.flush()  # supaya alt.id langsung ada

        # Buat Scores
        for cname, crit in criteria_map.items():
            value = row[col_index[cname]]
            score = Score(
                alternative_id=alt.id,
                criteria_id=crit.id,
                value=value
            )
            db.session.add(score)

        imported.append(name_alt)

    db.session.commit()

    return jsonify({"message": f"Import success"})
